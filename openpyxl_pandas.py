import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copyfile

template_file = 'Template.xlsx' # Has a header in row 1 already
output_file = 'Result.xlsx' # What we are saving the template as

# Copy Template.xlsx as Result.xlsx
copyfile(template_file, output_file)

# Read in the data to be pasted into the termplate
df = pd.read_csv('my_data.csv')

# Load the workbook and access the sheet we'll paste into
wb = load_workbook(output_file)
ws = wb.get_sheet_by_name('Existing Result Sheet')

# Selecting a cell in the header row before writing makes append()
#  start writing to the following line i.e. row 2
ws['A1']
# Write each row of the DataFrame
# In this case, I don't want to write the index (useless) or the header (already in the template)
for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

wb.save(output_file)


from openpyxl.utils.dataframe import dataframe_to_rows
rows = dataframe_to_rows(df)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         ws.cell(row=r_idx, column=c_idx, value=value)
