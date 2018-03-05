import openpyxl
import pandas as pd
import os

file_name = "./期货公司排名.xlsx"
# file_name = "./pandas_simple.xlsx"
writer = pd.ExcelWriter(file_name, engine='openpyxl')

if os.path.exists(file_name):
    book = openpyxl.load_workbook(file_name)
    writer.book = book
key = 'sheet3'
df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})
df.to_excel(writer, sheet_name=key)
writer.save()
writer.close()

#Write PNG file to existing worksheet
from openpyxl import Workbook
from openpyxl.drawing.image import Image
xfile = openpyxl.load_workbook('Things.xlsx')
sheet = xfile.get_sheet_by_name('Types of Cats')
img = Image('Typesofcats.png')
sheet.add_image(img, 'L6')
sheet1 = xfile.get_sheet_by_name('Types of Dogs')
img1 = Image('Typesofdogs.png')
sheet1.add_image(img1, 'I6')
sheet2 = xfile.get_sheet_by_name('Types of Pigs')
img2 = Image('Typesofpigs.png')
sheet2.add_image(img2, 'I6')
xfile.save('Things.xlsx')

import matplotlib.pyplot as plt
import openpyxl

# Your plot generation code here...
plt.savefig("myplot.png", dpi = 150)

wb = openpyxl.load_workbook('input.xlsx')
ws = wb.active

img = openpyxl.drawing.Image('myplot.png')
img.anchor(ws.cell('A1'))

ws.add_image(img)
wb.save('output.xlsx')